"""Export backtest results to CSV, Excel, and Markdown formats.

All string cells are sanitised to defuse spreadsheet formula injection. The
Markdown report includes a disclaimer, validation warnings first, headline
metrics, and per-segment breakdowns.
"""

from __future__ import annotations

import logging
from pathlib import Path

from claudetrade.backtest.engine import BacktestResult
from claudetrade.utils.text import sanitize_for_export

log = logging.getLogger(__name__)

try:
    import pandas as pd
except ImportError:
    pd = None  # type: ignore

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None  # type: ignore


def metrics_to_dataframe(metrics: dict) -> pd.DataFrame:
    """Convert a metrics dict to a DataFrame row.

    Args:
        metrics: A PerformanceMetrics instance's __dict__ or similar.

    Returns:
        A single-row DataFrame.
    """
    if pd is None:
        raise ImportError("pandas is required for metrics_to_dataframe")

    row = {}
    for key, value in metrics.items():
        if isinstance(value, (list, dict)):
            row[key] = str(value)
        else:
            row[key] = value

    return pd.DataFrame([row])


def trades_to_dataframe(result: BacktestResult) -> pd.DataFrame:
    """Convert trades to a DataFrame.

    Columns include entry/exit prices, P&L, holding period, MFE/MAE, and
    classification (win/loss/breakeven). All string columns are sanitised.

    Args:
        result: BacktestResult containing trades.

    Returns:
        A DataFrame with one row per trade.
    """
    if pd is None:
        raise ImportError("pandas is required for trades_to_dataframe")

    rows = []
    for trade in result.trades:
        rows.append({
            "trade_id": sanitize_for_export(trade.trade_id),
            "symbol": sanitize_for_export(trade.symbol),
            "strategy": sanitize_for_export(trade.strategy),
            "direction": sanitize_for_export(trade.direction.value),
            "entry_session": trade.entry_session,
            "entry_price": trade.entry_price,
            "shares": trade.shares,
            "stop_loss": trade.stop_loss,
            "exit_session": trade.exit_session,
            "exit_price": trade.exit_price,
            "exit_reason": sanitize_for_export(trade.exit_reason.value if trade.exit_reason else ""),
            "holding_days": trade.holding_days,
            "gross_pnl": trade.gross_pnl,
            "net_pnl": trade.net_pnl,
            "net_return_pct": trade.net_return_pct,
            "r_multiple": trade.r_multiple,
            "outcome": sanitize_for_export(trade.outcome().value if not trade.is_open else "open"),
            "mfe_pct": trade.mfe_pct,
            "mae_pct": trade.mae_pct,
            "mfe_r": trade.mfe_r,
            "mae_r": trade.mae_r,
            "sector": sanitize_for_export(trade.sector),
            "regime_at_entry": sanitize_for_export(trade.regime_at_entry.value),
        })

    return pd.DataFrame(rows)


def equity_to_dataframe(result: BacktestResult) -> pd.DataFrame:
    """Convert equity curve to a DataFrame.

    Columns include session, equity, cash, positions, exposure, and drawdown.

    Args:
        result: BacktestResult containing equity curve.

    Returns:
        A DataFrame with one row per session.
    """
    if pd is None:
        raise ImportError("pandas is required for equity_to_dataframe")

    rows = []
    for point in result.equity_curve:
        rows.append({
            "session": point.session,
            "equity": point.equity,
            "cash": point.cash,
            "open_positions": point.open_positions,
            "exposure_pct": point.exposure_pct,
            "portfolio_heat_pct": point.portfolio_heat_pct,
            "drawdown_pct": point.drawdown_pct,
        })

    return pd.DataFrame(rows)


def export_csv(result: BacktestResult, path: str | Path) -> None:
    """Export backtest results to CSV files.

    Creates three files in the target directory:
    - trades.csv: one row per trade
    - equity_curve.csv: daily marks
    - metrics.csv: headline metrics

    Args:
        result: BacktestResult to export.
        path: Target directory (created if missing).
    """
    if pd is None:
        raise ImportError("pandas is required for export_csv")

    path_obj = Path(path)
    path_obj.mkdir(parents=True, exist_ok=True)

    # Trades
    trades_df = trades_to_dataframe(result)
    trades_df.to_csv(path_obj / "trades.csv", index=False)
    log.info(f"Exported trades to {path_obj / 'trades.csv'}")

    # Equity curve
    equity_df = equity_to_dataframe(result)
    equity_df.to_csv(path_obj / "equity_curve.csv", index=False)
    log.info(f"Exported equity curve to {path_obj / 'equity_curve.csv'}")

    # Metrics
    metrics_df = metrics_to_dataframe(result.metrics)
    metrics_df.to_csv(path_obj / "metrics.csv", index=False)
    log.info(f"Exported metrics to {path_obj / 'metrics.csv'}")


def export_excel(result: BacktestResult, path: str | Path) -> None:
    """Export backtest results to a single Excel workbook.

    Sheets: trades, equity_curve, metrics, summary.

    Args:
        result: BacktestResult to export.
        path: Target file path (xlsx).
    """
    if Workbook is None:
        raise ImportError("openpyxl is required for export_excel")

    path_obj = Path(path)
    path_obj.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Trades sheet
    ws_trades = wb.create_sheet("trades")
    trades_df = trades_to_dataframe(result)
    for r_idx, row in enumerate(trades_df.itertuples(index=False), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_trades.cell(row=r_idx, column=c_idx, value=value)
            if isinstance(value, str):
                cell.value = sanitize_for_export(value)

    # Add headers
    for c_idx, col_name in enumerate(trades_df.columns, start=1):
        ws_trades.cell(row=1, column=c_idx, value=sanitize_for_export(col_name))
    ws_trades.insert_rows(1)

    # Equity curve sheet
    ws_equity = wb.create_sheet("equity_curve")
    equity_df = equity_to_dataframe(result)
    for r_idx, row in enumerate(equity_df.itertuples(index=False), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws_equity.cell(row=r_idx, column=c_idx, value=value)

    for c_idx, col_name in enumerate(equity_df.columns, start=1):
        ws_equity.cell(row=1, column=c_idx, value=col_name)
    ws_equity.insert_rows(1)

    # Metrics sheet
    ws_metrics = wb.create_sheet("metrics")
    metrics_df = metrics_to_dataframe(result.metrics)
    for c_idx, col_name in enumerate(metrics_df.columns, start=1):
        ws_metrics.cell(row=1, column=c_idx, value=sanitize_for_export(str(col_name)))
    for r_idx, row in enumerate(metrics_df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            if isinstance(value, str):
                value = sanitize_for_export(value)
            ws_metrics.cell(row=r_idx, column=c_idx, value=value)

    # Summary sheet
    ws_summary = wb.create_sheet("summary", 0)
    ws_summary.append(["Backtest Summary"])
    ws_summary.append([])
    ws_summary.append(["Run ID", result.run_id])
    ws_summary.append(["Start Date", result.start_session])
    ws_summary.append(["End Date", result.end_session])
    ws_summary.append(["Strategies", ", ".join(result.strategy_names)])
    ws_summary.append(["Universe Size", result.universe_size])
    ws_summary.append([])
    ws_summary.append(["Code Version", result.code_version])
    ws_summary.append(["Config Hash", result.config_hash])

    wb.save(path_obj)
    log.info(f"Exported Excel workbook to {path_obj}")


def render_markdown_report(result: BacktestResult) -> str:
    """Render a complete backtest report as Markdown.

    Includes:
    - Disclaimer header
    - Validation warnings (if any)
    - Headline metrics table
    - Per-segment breakdowns
    - Methodology note

    Args:
        result: BacktestResult to report.

    Returns:
        Markdown string.
    """
    lines = []

    # Disclaimer
    lines.append("# Backtest Report\n")
    lines.append(
        "⚠️ **Disclaimer**: This report shows historical simulation results only. "
        "Past performance does not guarantee future results. Strategy parameters "
        "were optimised on this data; out-of-sample performance may differ. "
        "See validation warnings below.\n"
    )

    # Validation warnings first
    if result.warnings:
        lines.append("## ⚠️ Validation Warnings\n")
        for warning in result.warnings:
            lines.append(f"- {warning}\n")
        lines.append("")

    # Metadata
    lines.append("## Run Information\n")
    lines.append(f"- **Run ID**: {result.run_id}\n")
    lines.append(f"- **Period**: {result.start_session} to {result.end_session}\n")
    lines.append(f"- **Strategies**: {', '.join(result.strategy_names)}\n")
    lines.append(f"- **Universe**: {result.universe_size} symbols\n")
    lines.append(f"- **Initial Capital**: ${result.config.initial_capital_usd:,.0f}\n")
    lines.append(f"- **Code Version**: {result.code_version}\n")
    lines.append(f"- **Config Hash**: {result.config_hash}\n")
    lines.append("")

    # Headline metrics
    from claudetrade.backtest.metrics import PerformanceMetrics
    metrics = PerformanceMetrics(**result.metrics)

    lines.append("## Headline Metrics\n")
    lines.append("| Metric | Value |\n")
    lines.append("|--------|-------|\n")

    lines.append(f"| Trades | {metrics.trade_count} |\n")
    lines.append(f"| Winning | {metrics.winning_trades} |\n")
    lines.append(f"| Losing | {metrics.losing_trades} |\n")
    lines.append(f"| Breakeven | {metrics.breakeven_trades} |\n")
    lines.append(f"| Win Rate | {100*metrics.win_rate:.1f}% |\n")

    if metrics.win_loss_ratio_is_degenerate:
        lines.append("| Win/Loss Ratio | ∞ (degenerate) |\n")
    else:
        lines.append(f"| Win/Loss Ratio | {metrics.win_loss_ratio:.2f} |\n")

    lines.append(f"| Expectancy | ${metrics.expectancy_dollars:,.2f} |\n")
    lines.append(f"| Profit Factor | {metrics.profit_factor:.2f} |\n")
    lines.append(f"| Avg Win | ${metrics.average_win:,.2f} |\n")
    lines.append(f"| Avg Loss | ${metrics.average_loss:,.2f} |\n")
    lines.append("")

    lines.append("## Risk & Return\n")
    lines.append("| Metric | Value |\n")
    lines.append("|--------|-------|\n")
    lines.append(f"| Total Return | {metrics.total_return_pct:.2f}% |\n")
    lines.append(f"| Annualised Return | {metrics.annualised_return_pct:.2f}% |\n")
    lines.append(f"| Max Drawdown | {metrics.max_drawdown_pct:.2f}% |\n")
    lines.append(f"| Max DD Duration | {metrics.max_drawdown_duration_days} days |\n")
    lines.append(f"| Sharpe Ratio | {metrics.sharpe:.2f} |\n")
    lines.append(f"| Sortino Ratio | {metrics.sortino:.2f} |\n")
    lines.append(f"| Avg Holding Days | {metrics.average_holding_days:.1f} |\n")
    lines.append(f"| Exposure | {metrics.exposure_pct:.1f}% |\n")
    lines.append("")

    # Segment metrics
    if result.segment_metrics:
        lines.append("## Segment Metrics\n")

        for dimension, segments in result.segment_metrics.items():
            if not segments:
                continue

            lines.append(f"\n### By {dimension}\n")
            lines.append("| Segment | Trades | W/L | Expectancy | Profit Factor |\n")
            lines.append("|---------|--------|-----|------------|---------------|\n")

            for segment_name, seg_metrics in sorted(segments.items()):
                seg_metrics_obj = (
                    seg_metrics
                    if isinstance(seg_metrics, PerformanceMetrics)
                    else PerformanceMetrics(**seg_metrics.__dict__ if hasattr(seg_metrics, "__dict__") else seg_metrics)
                )
                wl = (
                    "∞" if seg_metrics_obj.win_loss_ratio_is_degenerate
                    else f"{seg_metrics_obj.win_loss_ratio:.2f}"
                )
                lines.append(
                    f"| {sanitize_for_export(segment_name)} | "
                    f"{seg_metrics_obj.trade_count} | {wl} | "
                    f"${seg_metrics_obj.expectancy_dollars:,.0f} | "
                    f"{seg_metrics_obj.profit_factor:.2f} |\n"
                )

    # Methodology
    lines.append("\n## Methodology\n")
    lines.append(
        "- **Execution**: Daily bars, orders executed on next bar at configured reference price.\n"
    )
    lines.append(f"- **Execution Delay**: {result.config.execution_delay_bars} bar(s).\n")
    lines.append(
        "- **Costs**: Spread, slippage, and participation-cap fills per "
        "`claudetrade.backtest.costs`.\n"
    )
    lines.append(
        "- **Delisting**: Positions closed at close * 0.30 (recovery factor) "
        "with ExitReason.DELISTED.\n"
    )
    lines.append(
        "- **Intrabar Ambiguity**: Pessimistic policy (stop wins over target); "
        "see `claudetrade.backtest.execution`.\n"
    )
    lines.append("- **No Trade Left Open**: All positions force-closed at end-of-backtest.\n")

    return "".join(lines)


__all__ = [
    "equity_to_dataframe",
    "export_csv",
    "export_excel",
    "metrics_to_dataframe",
    "render_markdown_report",
    "trades_to_dataframe",
]
